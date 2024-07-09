#import necessary libraries

import openpyxl
import re

# Load the two Excel files
wb1 = openpyxl.load_workbook(r'validated_data.xlsx')
wb2 = openpyxl.load_workbook(r'final_data2.xlsx')

# Select the first sheet from each file
sheet1 = wb1.active
sheet2 = wb2.active

# Define a function to replace text and clean symbols
def replace_text(cell):
    if isinstance(cell.value, str):
        # Remove symbols and special characters
        cell.value = re.sub(r'[^\w\s]', '', cell.value)
        
        # Define regex pattern to match the specific phrases
        pattern1 = r"Im stopping the model because I couldnt find any brand name matching nan"
        pattern2 = r"Im stopping the model since I couldnt find any brand name matching nan"
        pattern3 = r"Nan"
        pattern4 = r'The corrected text is '
        
        # Use regex substitution to replace with an empty string
        cell.value = re.sub(pattern1, '', cell.value)
        cell.value = re.sub(pattern2, '', cell.value)
        cell.value = re.sub(pattern3, '', cell.value)
        cell.value = re.sub(pattern4, '', cell.value)

for ws in [sheet1, sheet2]:
    for row in ws.iter_rows(min_row=2, max_row=max(ws.max_row, sheet2.max_row), min_col=1, max_col=ws.max_column):
        for cell in row:
            replace_text(cell)
                       

# Specify the columns to compare (e.g. 'A', 'B', 'C' for columns A, B, and C)
columns_to_compare = ['A', 'B', 'E']

# Define a dictionary to map column letters to fill colors
column_colors = {
    'A': 'D8A3E1',  
    'B': 'C7E1A3',  
    'E': 'E1AEA3'   
}

for row1, row2 in zip(sheet1.iter_rows(), sheet2.iter_rows()):
    for cell1, cell2 in zip(row1, row2):
        column_index = cell1.column
        if column_index in [openpyxl.utils.column_index_from_string(column) for column in columns_to_compare]:
            # Ignore blank cells
            if cell1.value is None or cell2.value is None:
                continue

            '''# Ignore cells with errors
            if cell1.data_type == 'Do' or cell2.data_type == 'Do':
                continue'''

            # Convert values to strings before comparing
            if str(cell1.value) != str(cell2.value):
                # Highlight the cell in the second sheet with a color based on the column
                column_letter = openpyxl.utils.get_column_letter(column_index)
                fill_color = column_colors[column_letter]
                cell2.fill = openpyxl.styles.PatternFill(start_color=fill_color, end_color=fill_color, fill_type='solid')

# Save the changes to the second file
wb2.save('finalchecknew.xlsx')