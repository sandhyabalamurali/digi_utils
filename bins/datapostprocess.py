import os
from groq import Groq
from dotenv import load_dotenv
import pandas as pd
from openpyxl.styles import PatternFill

# Load the excel file, skipping the first row and using the second row as the header
data = pd.read_excel(r'data/input.xlsx')

# Drop the first row as it is the redundant header row
data_cleaned = data.drop([0,1]).reset_index(drop=True)

# Rename columns for easier reference
data_cleaned.columns = ['Equipment', 'Brand_Model', 'Capacity', 'Power_Rating_Watt', 'Qty_Nos', 'Usage_Hours', 'Working_Status']
data_cleaned

data.dtypes
len(data) #30 
len(data_cleaned) #24

# Remove empty rows (rows with all NaN values)
data_cleaned = data_cleaned.dropna(how='all')


# Identify and remove rows that contain only one non-null value
rows_to_remove = data_cleaned.apply(lambda x: x.count() == 1, axis=1)
data_cleaned = data_cleaned[~rows_to_remove]

# Reset index after cleaning
data_cleaned = data_cleaned.reset_index(drop=True)
print(data_cleaned)

# Define the check_int function
def check_int(value):
    try:
        return int(value)
    except ValueError:
        return value  # Return the original value if it cannot be converted to int
    
# Define the columns to be validated
# These are the indices of 'Power_Rating_Watt', 'Qty_Nos', and 'Usage_Hours'
columns_to_validate = [3, 4, 5]  

# Function to capitalize first letter and lowercase rest
def capitalize_first_letter(value):
    if isinstance(value, str):
        return value[:1].upper() + value[1:].lower()
    return value 

# Apply function to the last column
data_cleaned['Working_Status'] = data_cleaned['Working_Status'].apply(capitalize_first_letter)
print(data_cleaned['Working_Status'])
output_file = 'data/output/cleaned_data1.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    data_cleaned.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Define the yellow fill for highlighting
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Loop through the specified columns
    for col_index in columns_to_validate:
        for row in range(3, len(data_cleaned) + 2):  # Start from the second row to skip the header
            cell = worksheet.cell(row=row, column=col_index + 1)  # Adding 1 for 1-based index
            if cell.value is not None and cell.value != "":  # Check if the cell is not empty
                try:
                    # Attempt to convert to integer
                    int(cell.value)
                except ValueError:
                    # Apply yellow fill if value is not an integer
                    cell.fill = yellow_fill

# Load environment variables from .env file
load_dotenv(override=True)

# Retrieve the API key from environment variables
api_key = os.environ.get("GROQ_API_KEY")
if not api_key:
    raise ValueError("GROQ_API_KEY environment variable not set")

# Load the Excel file
df = pd.read_excel(r'data/output/cleaned_data1.xlsx')


def correct(text):
    try:
        client = Groq(api_key=api_key)
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    "content": f"Correct the following text and give the output in maximum 3 words only: {text}",
                }
            ],
            model="llama3-70b-8192",
            temperature=0.4,
            max_tokens=100,
            top_p=0.7,
            seed=10,
        )
        corrected_text = chat_completion.choices[0].message.content
        return corrected_text
    except Exception as e:
        print(f"Error: {e}")
        return text

def crtbrand(text):
    try:
        client = Groq(api_key=api_key)
        chat_completion = client.chat.completions.create(
            messages=[
                {
                    "role": "user",
                    # "content": f"Check the following brand name and return the output in double quotes without any messages, and if any error print 'value not found': {text}",
                    "content": f"Check the following brand name and return the correct brand in just a single word, and if you not found any values then stop the model: {text}"
                }
            ],
            model="llama3-8b-8192",
            temperature=0.4,
            top_p=0.7,
            seed=10,
        )
        corrected_text = chat_completion.choices[0].message.content
        return corrected_text
    except Exception as e:
        print(f"Error: {e}")
        return "value not found"

column_to_correct = 'Equipment'
column_to_correct2 = 'Brand_Model'

# Apply the correction function to the specified columns
df[column_to_correct] = df[column_to_correct].apply(correct)
df[column_to_correct2] = df[column_to_correct2].apply(crtbrand)

# Save the updated DataFrame to a new Excel file
df.to_excel(r'data/output/spell_output.xlsx', index=False)
data_clean= pd.read_excel('data/output/spell_output.xlsx', header=0)

output_file = 'Final_data.xlsx'
with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
    data_clean.to_excel(writer, index=False, sheet_name='Sheet1')
    workbook = writer.book
    worksheet = writer.sheets['Sheet1']
    
    # Define the yellow fill for highlighting
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    
    # Loop through the specified columns
    for col_index in columns_to_validate:
        for row in range(3, len(data_clean) + 2):  # Start from the second row to skip the header
            cell = worksheet.cell(row=row, column=col_index + 1)  # Adding 1 for 1-based index
            if cell.value is not None and cell.value != "":  # Check if the cell is not empty
                try:
                    # Attempt to convert to integer
                    int(cell.value)
                except ValueError:
                    # Apply yellow fill if value is not an integer
                    cell.fill = yellow_fill